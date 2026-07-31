"""
Extrae el contenido del Excel "FTA Gen (Hito1)" a JSON, normalizando las
columnas (que cambian de hoja en hoja) y los IDs (que Excel convirtió en fechas).
"""
import openpyxl, datetime, json, re

SRC = '/Users/johanpina/Downloads/FTA Gen (Hito1) Contenido_Desarrollo  .xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

def fmt(v):
    if v is None: return ''
    if isinstance(v, datetime.datetime):
        return f'{v.month}.{v.day}'          # "1.4", "6.2" leídos como fechas
    return str(v).replace('\r', '').strip()

VACIO = {'', '-', '--', 'se mantiene'}

def limpio(s):
    """Normaliza una celda de contenido; '-' y 'Se mantiene' son marcadores, no texto."""
    s = s.lstrip('- ').strip()
    return '' if s.lower() in VACIO else s

def vigente(actual, nuevo):
    """
    Qué texto rige. La columna 'Contenido nuevo' se usa de dos formas en el
    documento: para reemplazar contenido que no existía ('-' en 'actual'), y para
    anotar comentarios sobre contenido que sí existe. Por eso 'actual' manda
    cuando trae algo: en ninguna fila del documento el 'nuevo' reemplaza de
    verdad a un 'actual' con contenido (verificado sobre las 12 hojas).
    """
    return actual or nuevo

def colmap(ws):
    """Las hojas no comparten orden de columnas: se mapean por encabezado."""
    m = {}
    for i, c in enumerate(ws[1]):
        h = fmt(c.value).lower()
        if not h: continue
        if h.startswith('id') or h.startswith('vision'): m['id'] = i
        elif h == 'bloque': m['bloque'] = i
        elif h == 'tipo': m['tipo'] = i
        elif h.startswith('contenido actual'): m['actual'] = i
        elif h.startswith('contenido nuevo'): m['nuevo'] = i
        elif h.startswith('condición'): m['condicion'] = i
        elif h.startswith('comentario'): m['comentario'] = i
        elif h.startswith('estado'): m['estado'] = i
        elif h.startswith('nueva hito'): m['hito'] = i
    return m

dims = []
for ws in wb.worksheets:
    if ws.title.startswith('Encuesta'): continue
    m = colmap(ws)
    preguntas, actual = [], None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        g = lambda k: fmt(row[m[k]].value) if k in m and m[k] < len(row) else ''
        tipo = g('tipo').lower()
        if not any(fmt(c.value) for c in row): continue

        if tipo == 'label':
            actual = {
                'id': g('id'),
                'bloque': g('bloque'),
                'label_actual': limpio(g('actual')),
                'label_nuevo': limpio(g('nuevo')),
                'obligatoria': g('condicion'),
                'comentario': g('comentario'),
                'estado': g('estado'),
                'hito': g('hito'),
                'tooltip': '', 'placeholder': '', 'opciones': '', 'alternativas': '',
            }
            actual['texto'] = vigente(actual['label_actual'], actual['label_nuevo'])
            preguntas.append(actual)
        elif actual is not None:
            texto = vigente(limpio(g('actual')), limpio(g('nuevo')))
            if tipo == 'tooltip': actual['tooltip'] = texto
            elif tipo == 'placeholder': actual['placeholder'] = texto
            elif tipo == 'option': actual['opciones'] = texto
            elif tipo.startswith('alternativa'): actual['alternativas'] = texto
    dims.append({'hoja': ws.title.strip(), 'preguntas': preguntas})

# ── Encuesta ────────────────────────────────────────────────────────
ws = wb['Encuesta facilida de uso']
enc, actual = [], None
for row in ws.iter_rows(min_row=2, max_row=40):
    idv, tipo, nuevo, cond = fmt(row[0].value), fmt(row[1].value).lower(), fmt(row[3].value), fmt(row[4].value)
    if not idv: continue
    if tipo == 'pregunta':
        actual = {'id': idv, 'pregunta': nuevo, 'obligatoria': cond,
                  'escala': '', 'tooltip': '', 'placeholder': ''}
        enc.append(actual)
    elif actual is not None:
        if tipo == 'option': actual['escala'] = nuevo
        elif tipo == 'tooltip': actual['tooltip'] = nuevo
        elif tipo == 'placeholder': actual['placeholder'] = nuevo

out = {'dimensiones': dims, 'encuesta': enc}
dest = '/Users/johanpina/dev/Herramientas_Goblab/herramienta_transparencia_modelo/docs/fta-gen/contenido-hito1.json'
json.dump(out, open(dest, 'w'), ensure_ascii=False, indent=2)
print('dimensiones:', len(dims), '| preguntas:', sum(len(d['preguntas']) for d in dims), '| encuesta:', len(enc))
